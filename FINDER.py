import openpyxl

# Open the workbook
workbook = openpyxl.load_workbook('PayerDataNEW1.xlsx')

# Get the sheet names
sheet_names = workbook.sheetnames

# Get the first sheet
first_sheet = workbook[sheet_names[0]]

# Iterate over the rows in columns A, B, and C of the first sheet
for row in range(1, first_sheet.max_row + 1):
    short_name = first_sheet.cell(row=row, column=1).value
    payer_name = first_sheet.cell(row=row, column=2).value
    formats = first_sheet.cell(row=row, column=3).value
    value_to_copy = None

    # Iterate over the other sheets
    for sheet_name in sheet_names[1:]:
        sheet = workbook[sheet_name]

        # Iterate over the rows in columns A, B, and C of the current sheet
        for other_row in range(1, sheet.max_row + 1):
            other_short_name = sheet.cell(row=other_row, column=1).value
            other_payer_name = sheet.cell(row=other_row, column=2).value
            other_formats = sheet.cell(row=other_row, column=3).value

            # If any one of the values match, copy the value from column D and break out of the loop
            if short_name == other_short_name or payer_name == other_payer_name or formats == other_formats:
                first_sheet.cell(row=row, column=4).value = sheet.cell(row=other_row, column=4).value
                value_to_copy = sheet.cell(row=other_row, column=4).value
                break

            if short_name == other_short_name or formats == other_formats:
                first_sheet.cell(row=row, column=4).value = sheet.cell(row=other_row, column=4).value
                value_to_copy = sheet.cell(row=other_row, column=4).value
                break

        # If we found a match, break out of the inner loop
        if value_to_copy is not None:
            break

# Save the workbook
workbook.save('PayerDataNEW.xlsx')
